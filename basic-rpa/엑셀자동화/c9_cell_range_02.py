# 전체 행, 열 셀값 가져오기

from openpyxl import Workbook
from openpyxl import load_workbook
from random import *

wb = load_workbook("sample.xlsx") # # sample.xlsx 파일에서 wb를 불러옴
ws = wb.active

# 전체 rows 출력
# print("--- rows ---")
# print(ws.rows)
# print(tuple(ws.rows)) # 전체 셀 정보를 튜플 형태로 가져옴
# for row in tuple(ws.rows):
#     print(row[0].value)
# for row in ws.iter_rows(): 
#     print(row[0].value)


# 전체 columns
# print("--- columns ---")
# print(ws.columns)
# print(tuple(ws.columns)) # 전체 셀 정보를 튜플 형태로 가져옴
# for col in tuple(ws.columns):
#     print(col[0].value)
# for col in ws.iter_cols():
#     print(col[0].value, end=" ")

print("-------------------------------------")
for row in ws.iter_rows(): 
    for i in range(0, 5):
        print(row[i].value, end="\t")
    print()
print("-------------------------------------")

for row in ws.iter_rows(min_row=1, max_row=3):
    print(row[2].value, end="\t")

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)

wb.save("sample.xlsx")

