# 조건으로 셀값 찾기/수정

from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):
    # 번호, 영어, 수학
    if int(row[2].value) > 80:
        print(row[0].value, "번 학생은 영어 천재")

for row in ws.iter_rows(max_row=3):
    for cell in row:
        print(cell.value)
        if cell.value == "영어":
            cell.value = "외국어"

wb.save("sample_modified.xlsx")