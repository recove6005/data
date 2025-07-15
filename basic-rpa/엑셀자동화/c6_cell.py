# Section2. 엑셀 자동화
# Chapter 6. 셀 기본

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "MySheet"

# 셀에 값 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1 셀의 정보 출력
print(ws["A1"].value) # A1 셀의 값 출력
print(ws["A10"].value) # 셀 값이 없을 땐 'None' 출력

# row = 1, 2, 3, ... 
# column = A(1), B(2), C(3), ...
print(ws.cell(row=1, column=1).value) # = ws["A1"].value
print(ws.cell(row=1, column=2).value) # = ws["B1"].value

c = ws.cell(row=1, column=3, value=10) # ws["C1"].value = 10
print(c.value) # ws["C1"]의 값 출력


# 반복문을 이용해 랜덤 숫자 채우기
from random import *

for row in range(1, 11):
    for col in range(1, 11):
        ws.cell(row=row, column=col, value=randint(0, 100))

wb.save("sample.xlsx")