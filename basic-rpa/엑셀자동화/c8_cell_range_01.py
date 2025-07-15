# 셀값 가져오기

from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "국어", "영어", "수학", "코딩"]) # A1, B1, C1, D1, E1

# 10개 데이터 넣기
for i in range(1, 11): 
    ws.append([i, randint(0, 100), randint(0, 100), randint(0, 100), randint(0, 100)])

# 국어 colum 값들만 가져오기
col_B = ws["B"]
for cell in col_B:
    print(cell.value, end=", ")
print()

# 국어~영어(B~C) colum 셀값 가져오기
col_range = ws["B:C"]
for cols in col_range:
    for cell in cols:
        print(cell.value, end=", ")
    print()

# 2번째 줄에서 6번째 줄까지 셀값 가져오기
row_range = ws[2:6] 
for rows in row_range:
    for cell in rows:
        print(cell.value, end=", ")
    print()

# 2번째 줄에서 마지막줄까지 셀값 가져오기
row_range = ws[2:ws.max_row]
for rows in row_range:
    for cell in rows:
        print(cell.value, end=", ")
        print(cell.coordinate, end=", ")
    print()

# 2번째 줄에서 마지막줄까지 셀이름 가져오기
row_range = ws[2:ws.max_row]
for rows in row_range:
    for cell in rows:
        print(cell.coordinate, end=", ")
    print()

from openpyxl.utils.cell import coordinate_from_string
row_range = ws[2:ws.max_row]
for rows in row_range:
    for cell in rows:
        xy = coordinate_from_string(cell.coordinate) # 해당 셀의 셀이름 가져오기
        # print(xy, end=" ")
        print(xy[0], end= "*") # A
        print(xy[1], end="- ") # 1

wb.save("sample.xlsx")

