from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 셀 병합
ws.merge_cells("B2:D2") # B2~D2 셀병합
ws["B2"].value = "Merged Cell"
wb.save("sample_merge.xlsx")

# 셀 분리
ws.unmerge_cells("B2:D2") # B2~D2 병합 해제
wb.save("sample_unmerge.xlsx")

