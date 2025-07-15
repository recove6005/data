from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 행열 셀 삽입
ws.insert_rows(8) # 8번 줄에 빈 row 삽입
ws.insert_rows(8, 5) # 8번 줄에서 5개의 빈 row 삽입
ws.insert_cols(2) # B번째 열이 비워짐 (새로운 빈 열이 추가)
ws.insert_cols(2, 3) # B번째 열에서 3개의 빈 column 삽입

# 행열 셀 삭제
ws.delete_rows(8) # 8행 데이터 삭제
ws.delete_rows(8, 3) # 8행부터 3개 행 데이터 삭제
ws.delete_cols(2) # B열 데이터 삭제
ws.delete_cols(2, 3) # B열부터 2개열 데이터 삭제

# 행열 셀 이동
ws.move_range("B1:C11", rows=0, cols=1) # B1~C11 영역을 오른쪽으로 1칸 이동
ws["B1"].value = " 국어" # 이동 후 비워진 B1 셀에 값 입력
ws.move_range("B1:C11", rows=3, cols=1) # B1~C11 영역을 오른쪽으로 1칸, 아래쪽으로 3칸 이동
ws.move_range("C1:C11", rows=5, cols= 0) # C열을 아래쪽으로 5칸 이동
ws.save("sample.xlsx")