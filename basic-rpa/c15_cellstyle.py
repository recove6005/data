from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
wb = load_workbook("sample.xlsx")
ws = wb.active

a1 = ws["A1"] # 번호
b1 = ws["B1"] # 영어
c1 = ws["C1"] # 수학

# A 열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5 

# 1행 높이를 50으로 설정
ws.row_dimensions[1].height = 50

# 폰트 스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True)
b1.font = Font(color="CC33FF", name="Arial", strike=True)
c1.font = Font(color="0000FF", size=20, underline="single")

# 테두리 적용
# thin, medium, thick, dashed, dotted
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
bold_border = Border(left=Side(style="thick"), right=Side(style="thick"), top=Side(style="thick"), bottom=Side(style="thick"))
dot_border = Border(left=Side(style="dotted"), right=Side(style="dotted"), top=Side(style="dotted"), bottom=Side(style="dotted"))
a1.border = thin_border
b1.border = bold_border
c1.border = dot_border

# 셀 색상 변경




wb.save("sample_style.xlsx")