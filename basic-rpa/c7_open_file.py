from openpyxl import load_workbook # 파일 불러오기

wb = load_workbook("sample.xlsx") # # sample.xlsx 파일에서 wb를 불러옴
ws = wb.active # 활성화된 Sheet

# cell 데이터 불러오기
for row in range(1, 11):
    for col in range(1, 11):
        print(ws.cell(row=row, column=col).value, end=" ")
    print()

# cell 갯수를 모를 때
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        print(ws.cell(row=row, column=col).value, end=" ")
    print()

