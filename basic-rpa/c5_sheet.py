# Section2. 엑셀 자동화
# Chapter 5. 시트
from openpyxl import Workbook

wb = Workbook()
wb.active # Sheet

ws = wb.create_sheet() # 새로운 시트 생성
ws.title = "MySheet1" # 새로 생성한 Sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # RGB 형태로 값을 넣어주변 새로 생성한 시트 탭의 색상이 변경됨 (Colors RGB)

ws1 = wb.create_sheet("MySheet2")
ws2 = wb.create_sheet("MySheet3", 2) # 2번째 index에 Sheet 생성

# 모든 sheet의 이름 확인
print(wb.sheetnames) 

# Sheet 복사
mysheet3 = wb["MySheet3"] # MySheet3을 딕셔너리 형태로 접근
mysheet3["A1"] = "Test" # A1 셀에 "Test" 입력
target = wb.copy_worksheet(mysheet3) # target에 MySheet3 시트를 복사
target.title = "Copied Sheet" # 복사한 시트 이름 변경

wb.save("sample.xlsx") # 시트 저장